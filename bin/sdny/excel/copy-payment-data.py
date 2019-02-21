#!/usr/bin/env python3

# import sys
from xlrd import open_workbook,xldate_as_tuple
import openpyxl, json, sys, re
from datetime import date
import argparse

parser = argparse.ArgumentParser(description='copies our Excel payment-data into their spreadsheet')
parser.add_argument("--spreadsheet", help="path to spreadsheet containing our data to parse", required=True)
#parser.add_argument("--name-of-worksheet",help="name of worksheet to parse", required=True)
parser.add_argument("--target-spreadsheet",help="path to their spreadsheet",required=True)
parser.add_argument("--target-worksheet",help="name worksheet in which to enter data",required=True,
    choices=["FY17","FY18"])

args = parser.parse_args()
our_wb_file = args.spreadsheet
#worksheet = args.name_of_worksheet
target_spreadsheet = args.target_spreadsheet
target_worksheet = args.target_worksheet


def get_fy_and_quarter(date_obj):
    month = date_obj.month
    year  = date_obj.year
    # default
    fy = year;
    if month >= 10:
        fy = year + 1
        q  = 1
    elif month in range(1,4):
        q = 2
    elif month in range(4,7):
        q = 3
    elif month in range(7,10):
        q = 4

    return(str(fy)[2:], str(q))

def estimate_days(money):
    if (money < 418):
        return 0.5
    return 1.0

json_data_file = "data/{}.interp-languages.json".format(target_worksheet)

their_wb = openpyxl.load_workbook(target_spreadsheet)
if (target_worksheet == "FY17"):
    their_wb.active = 0;
else:
    their_wb.active = 1;

their_sheet = their_wb.get_active_sheet();

our_wb = open_workbook(our_wb_file)
sheet = our_wb.sheet_by_index(0)

# how are we doing?
print("using their {}, wksheet {}; our file is {}".format(
    target_spreadsheet, target_worksheet, our_wb_file
));
n = 2;
interpreters = json.loads(open(json_data_file).read())

for i in range(1,sheet.nrows):

    name = sheet.cell(i,0).value.strip()
    language = sheet.cell(i,1).value.strip()
    classification = None
    if language == "Spanish":
        continue
    if name in interpreters and language in interpreters[name]["languages"]:
        classification = interpreters[name]["languages"][language].upper()
    else:
        print("WARNING: can't determine classification for {} in {}".format(name,language))
    if classification and not classification == "LS":
        print("{} thought to be {} in {}, skipping record".format(name,classification,language))
        continue
    date_value = sheet.cell(i,2).value
    if not date_value:
        print("WARNING: no event-date value found in row {}".format(i))
        continue
    event_date_tuple = xldate_as_tuple(int(date_value),our_wb.datemode)
    event_date = date(*event_date_tuple[:3]).strftime("%m/%d/%Y")
    try:
        sent_to_judge_tuple = xldate_as_tuple(int(sheet.cell(i,18).value),our_wb.datemode)
        date_sent_to_judge = date(*(sent_to_judge_tuple[:3]))
        fy, q = get_fy_and_quarter(date_sent_to_judge)
        fy_and_quarter = "FY{} Q{}".format(fy, q)
    except:
        print("WARNING: can't parse date-sent-to-judge at row {}".format(i))
        continue
    try:
        docket = sheet.cell(i,8).value.strip().upper()
    except:
        print("WARNING: can't parse docket {} at row {}".format(sheet.cell(i,8).value,i))
        docket = None
    if not docket:
        docket = ""
        case_type = ""
    elif "CR" in docket:
        case_type = "Criminal"
    elif (re.search("CI?V",docket) != None):
        if (language == "ASL"):
            continue # so they tell us
        case_type = "Civil"
    if (not case_type):
        print("WARNING: can't infer case type from docket at row {}".format(i))
    event_type = sheet.cell(i,10).value.lower()
    if sheet.cell(i,11).value:
        money = sheet.cell(i,11).value # in-court
    elif sheet.cell(i,14).value:
        money = sheet.cell(i,14).value # out-of-court
    elif sheet.cell(i,17).value:
        money = sheet.cell(i,17).value # last-ditch, grand total
    else:
        print("WARNING: no money found at row {}, skipping".format(i))
        continue

    their_sheet["A{}".format(n)] = fy_and_quarter
    their_sheet["B{}".format(n)] = "2nd"
    their_sheet["C{}".format(n)] = "NYS"
    their_sheet["D{}".format(n)] = event_date
    their_sheet["E{}".format(n)] = case_type
    their_sheet["G{}".format(n)] = "LS"
    their_sheet["H{}".format(n)] = language
    their_sheet["F{}".format(n)] = event_type
    their_sheet["I{}".format(n)] = name
    their_sheet["N{}".format(n)] = money
    their_sheet["M{}".format(n)] = estimate_days(money)
    their_sheet["O{}".format(n)] = "Paula Gold"
    their_sheet["P{}".format(n)] = "Chief Interpreter"
    n += 1;

their_wb.save("data/excel/bullshit-table.xlsx")
their_wb.close()

print("All done inserting {} rows of payment data. You're welcome".format(n))
